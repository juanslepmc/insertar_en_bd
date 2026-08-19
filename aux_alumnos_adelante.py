import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

# RUTA DE LOS ARCHIVOS
FILE_A = 'excel/alumnos/alumnos_actual.xlsx'
FILE_B = 'excel/alumnos/Consolidado_sin_duplicados_agosto.xlsx'
FILE_OUTPUT = 'excel/alumnos/resultado_comparacion_alumnos.xlsx'

def normalizar_rut(cuerpo, dv=None):
    """
    Limpia y concatena el RUT eliminando puntos y espacios.
    Devuelve formato estándar: 12345678-K
    """
    if pd.isna(cuerpo):
        return ""
    cuerpo_str = str(cuerpo).strip()
    if cuerpo_str.endswith('.0'):
        cuerpo_str = cuerpo_str[:-2]
    cuerpo_str = cuerpo_str.replace('.', '').replace(' ', '').replace('-', '')
    
    if dv is not None:
        if pd.isna(dv):
            return ""
        dv_str = str(dv).strip().upper()
        if dv_str.endswith('.0'):
            dv_str = dv_str[:-2]
        dv_str = dv_str.replace('.', '').replace(' ', '')
        return f"{cuerpo_str}-{dv_str}" if cuerpo_str else ""
    else:
        if '-' in cuerpo_str:
            parts = cuerpo_str.rsplit('-', 1)
            return f"{parts[0]}-{parts[1].upper()}"
        elif len(cuerpo_str) > 1:
            return f"{cuerpo_str[:-1]}-{cuerpo_str[-1].upper()}"
        return cuerpo_str.upper()

def normalizar_genero(val):
    """
    Transforma 'F' -> 'Femenino', 'M' -> 'Masculino'.
    Conserva valores descriptivos si ya vienen completos.
    """
    if pd.isna(val):
        return None
    val_str = str(val).strip().upper()
    if val_str in ['F', 'FEMENINO']:
        return 'Femenino'
    elif val_str in ['M', 'MASCULINO']:
        return 'Masculino'
    return str(val).strip()

def normalizar_fecha(val):
    """
    Convierte fechas al formato estándar de MySQL (YYYY-MM-DD).
    """
    if pd.isna(val):
        return None
    dt = pd.to_datetime(val, errors='coerce')
    return dt.strftime('%Y-%m-%d') if pd.notna(dt) else None

def procesar_comparacion_alumnos():
    print("Cargando archivos...")
    df_a = pd.read_excel(FILE_A)
    df_b = pd.read_excel(FILE_B, sheet_name='Sheet1')

    # 1. Normalización de RUTs
    df_a['rut_norm'] = df_a['rut'].apply(lambda x: normalizar_rut(x))
    df_b['rut_norm'] = df_b.apply(lambda r: normalizar_rut(r['Run'], r['Dígito Ver.']), axis=1)

    # 2. Detección DINÁMICA de columnas en Archivo A
    col_genero_a = next((col for col in ['genero', 'sexo', 'Genero', 'Sexo'] if col in df_a.columns), None)
    col_fecha_a = next((col for col in ['fecha_nacimiento', 'Fecha Nacimiento', 'fecha_nac', 'Fecha_Nacimiento'] if col in df_a.columns), None)

    if col_genero_a:
        print(f"-> Columna de género detectada en Archivo A: '{col_genero_a}'")
        df_a['genero_fmt'] = df_a[col_genero_a].apply(normalizar_genero)
    else:
        print("-> Archivo A no posee columna de género. Se asignará None.")
        df_a['genero_fmt'] = None

    if col_fecha_a:
        print(f"-> Columna de fecha de nacimiento detectada en Archivo A: '{col_fecha_a}'")
        df_a['fecha_nac_fmt'] = df_a[col_fecha_a].apply(normalizar_fecha)
    else:
        print("-> Archivo A no posee columna de fecha de nacimiento. Se asignará None.")
        df_a['fecha_nac_fmt'] = None

    # 3. Procesamiento en Archivo B
    df_b['genero_fmt'] = df_b['Genero'].apply(normalizar_genero)
    df_b['fecha_nac_fmt'] = df_b['Fecha Nacimiento'].apply(normalizar_fecha)

    # Esquema de columnas definitivo para MySQL
    cols_finales = ['rut', 'nombres', 'apellido_paterno', 'apellido_materno', 'comuna', 'genero', 'fecha_nacimiento', 'estado']

    # Mapeo estructurado B
    df_b_mapped = pd.DataFrame({
        'rut': df_b['rut_norm'],
        'nombres': df_b['Nombres'],
        'apellido_paterno': df_b['Apellido Paterno'],
        'apellido_materno': df_b['Apellido Materno'],
        'comuna': df_b['Comuna Residencia'],
        'genero': df_b['genero_fmt'],
        'fecha_nacimiento': df_b['fecha_nac_fmt'],
        'estado': 'Activo',
        'rut_norm': df_b['rut_norm']
    })

    # Mapeo estructurado A
    df_a_mapped = pd.DataFrame({
        'rut': df_a['rut_norm'],
        'nombres': df_a['nombres'],
        'apellido_paterno': df_a['apellido_paterno'],
        'apellido_materno': df_a['apellido_materno'],
        'comuna': df_a['comuna'],
        'genero': df_a['genero_fmt'],
        'fecha_nacimiento': df_a['fecha_nac_fmt'],
        'estado': df_a['estado'],
        'rut_norm': df_a['rut_norm']
    })

    ruts_a = set(df_a['rut_norm'])
    ruts_b = set(df_b['rut_norm'])

    # 4. Separación por Escenarios
    # Escenario 1: Match (Se prefiere la versión actualizada de B)
    df_match = df_b_mapped[df_b_mapped['rut_norm'].isin(ruts_a)][cols_finales].copy()

    # Escenario 2: Nuevos en B (Activos)
    df_nuevos = df_b_mapped[~df_b_mapped['rut_norm'].isin(ruts_a)][cols_finales].copy()
    df_nuevos['estado'] = 'Activo'

    # Escenario 3: Solo en A (Inactivos)
    df_bajas = df_a_mapped[~df_a_mapped['rut_norm'].isin(ruts_b)][cols_finales].copy()
    df_bajas['estado'] = 'Inactivo'

    # 5. Generación de Excel con Estilos
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hojas = [
        ('Coincidencias (Match)', df_match, '1F4E78'),
        ('Nuevos en B (Activos)', df_nuevos, '276A3C'),
        ('Bajas de A (Inactivos)', df_bajas, '842323')
    ]

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    for sheet_title, df_data, fill_color in hojas:
        ws = wb.create_sheet(title=sheet_title)
        
        # Encabezado
        ws.append(cols_finales)
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        
        for col_idx in range(1, len(cols_finales) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Filas de datos
        fill_even = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for row_idx, row_values in enumerate(df_data.values, start=2):
            ws.append(list(row_values))
            row_fill = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx in range(1, len(cols_finales) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = row_fill
                c.border = thin_border
                c.font = Font(name='Calibri', size=10)
                col_name = cols_finales[col_idx - 1]
                if col_name in ['rut', 'genero', 'fecha_nacimiento', 'estado']:
                    c.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center')
                    
        # Autoajuste de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.views.sheetView[0].showGridLines = True

    # 6. Guardado seguro
    try:
        wb.save(FILE_OUTPUT)
        print(f"\n✅ Archivo generado exitosamente: {FILE_OUTPUT}")
    except PermissionError:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        alt_output = FILE_OUTPUT.replace(".xlsx", f"_{timestamp}.xlsx")
        print(f"\n⚠️ El archivo '{FILE_OUTPUT}' estaba abierto. Guardando en: {alt_output}")
        wb.save(alt_output)

    print(f"  - Coincidencias (Match): {len(df_match)} registros")
    print(f"  - Nuevos (Activos): {len(df_nuevos)} registros")
    print(f"  - Bajas (Inactivos): {len(df_bajas)} registros")

if __name__ == '__main__':
    procesar_comparacion_alumnos()