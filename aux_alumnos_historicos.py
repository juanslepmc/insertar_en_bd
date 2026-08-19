import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

# RUTA DE LOS ARCHIVOS
FILE_A = 'excel/alumnos/alumnos_actual.xlsx'  # Base de Datos actual (Dic 2025 - Ago 2026)
FILE_B = 'excel/alumnos/Consolidado_sin_duplicados_nov_2025.xlsx'  # Archivo nuevo histórico
FILE_OUTPUT = 'excel/alumnos/resultado_comparacion_historico.xlsx'

def normalizar_rut(cuerpo, dv=None):
    if pd.isna(cuerpo):
        return ""
    cuerpo_str = str(cuerpo).strip()
    if cuerpo_str.endswith('.0'):
        cuerpo_str = cuerpo_str[:-2]
    cuerpo_str = cuerpo_str.replace('.', '').replace(' ', '').replace('-', '')
    
    if dv is not None:
        if pd.isna(dv):
            return ""
        dv_str = str(dv).strip().upper()
        if dv_str.endswith('.0'):
            dv_str = dv_str[:-2]
        dv_str = dv_str.replace('.', '').replace(' ', '')
        return f"{cuerpo_str}-{dv_str}" if cuerpo_str else ""
    else:
        if '-' in cuerpo_str:
            parts = cuerpo_str.rsplit('-', 1)
            return f"{parts[0]}-{parts[1].upper()}"
        elif len(cuerpo_str) > 1:
            return f"{cuerpo_str[:-1]}-{cuerpo_str[-1].upper()}"
        return cuerpo_str.upper()

def normalizar_genero(val):
    if pd.isna(val):
        return None
    val_str = str(val).strip().upper()
    if val_str in ['F', 'FEMENINO']:
        return 'Femenino'
    elif val_str in ['M', 'MASCULINO']:
        return 'Masculino'
    return str(val).strip()

def normalizar_fecha(val):
    if pd.isna(val):
        return None
    dt = pd.to_datetime(val, errors='coerce')
    return dt.strftime('%Y-%m-%d') if pd.notna(dt) else None

def procesar_comparacion_historico():
    print("🚀 Cargando archivos para comparación histórica...")
    df_a = pd.read_excel(FILE_A)
    df_b = pd.read_excel(FILE_B)

    # 1. Normalización de RUTs
    df_a['rut_norm'] = df_a['rut'].apply(lambda x: normalizar_rut(x))
    
    # Detección dinámica de columna de RUT en B
    if 'Run' in df_b.columns and 'Dígito Ver.' in df_b.columns:
        df_b['rut_norm'] = df_b.apply(lambda r: normalizar_rut(r['Run'], r['Dígito Ver.']), axis=1)
    elif 'rut' in df_b.columns:
        df_b['rut_norm'] = df_b['rut'].apply(lambda x: normalizar_rut(x))
    else:
        raise ValueError("No se encontró columna de RUT/RUN válida en Archivo B.")

    # 2. Detección DINÁMICA de columnas de género y fecha en A y B
    col_gen_a = next((c for c in ['genero', 'sexo', 'Genero', 'Sexo'] if c in df_a.columns), None)
    col_fec_a = next((c for c in ['fecha_nacimiento', 'Fecha Nacimiento', 'fecha_nac'] if c in df_a.columns), None)
    col_gen_b = next((c for c in ['genero', 'sexo', 'Genero', 'Sexo'] if c in df_b.columns), None)
    col_fec_b = next((c for c in ['fecha_nacimiento', 'Fecha Nacimiento', 'fecha_nac'] if c in df_b.columns), None)

    df_a['genero_fmt'] = df_a[col_gen_a].apply(normalizar_genero) if col_gen_a else None
    df_a['fecha_nac_fmt'] = df_a[col_fec_a].apply(normalizar_fecha) if col_fec_a else None
    
    df_b['genero_fmt'] = df_b[col_gen_b].apply(normalizar_genero) if col_gen_b else None
    df_b['fecha_nac_fmt'] = df_b[col_fec_b].apply(normalizar_fecha) if col_fec_b else None

    cols_finales = ['rut', 'nombres', 'apellido_paterno', 'apellido_materno', 'comuna', 'genero', 'fecha_nacimiento', 'estado']

    # Mapeo B (Histórico)
    col_nom_b = next((c for c in ['Nombres', 'nombres', 'nombre'] if c in df_b.columns), None)
    col_pat_b = next((c for c in ['Apellido Paterno', 'apellido_paterno', 'paterno'] if c in df_b.columns), None)
    col_mat_b = next((c for c in ['Apellido Materno', 'apellido_materno', 'materno'] if c in df_b.columns), None)
    col_com_b = next((c for c in ['Comuna Residencia', 'comuna', 'comuna_residencia'] if c in df_b.columns), None)

    df_b_mapped = pd.DataFrame({
        'rut': df_b['rut_norm'],
        'nombres': df_b[col_nom_b] if col_nom_b else "Sin información",
        'apellido_paterno': df_b[col_pat_b] if col_pat_b else "Sin información",
        'apellido_materno': df_b[col_mat_b] if col_mat_b else None,
        'comuna': df_b[col_com_b] if col_com_b else "Sin información",
        'genero': df_b['genero_fmt'],
        'fecha_nacimiento': df_b['fecha_nac_fmt'],
        'rut_norm': df_b['rut_norm']
    })

    # Mapeo A (Actual)
    df_a_mapped = pd.DataFrame({
        'rut': df_a['rut_norm'],
        'nombres': df_a['nombres'],
        'apellido_paterno': df_a['apellido_paterno'],
        'apellido_materno': df_a['apellido_materno'],
        'comuna': df_a['comuna'],
        'genero': df_a['genero_fmt'],
        'fecha_nacimiento': df_a['fecha_nac_fmt'],
        'estado': df_a['estado'],
        'rut_norm': df_a['rut_norm']
    })

    ruts_a = set(df_a['rut_norm'])
    ruts_b = set(df_b['rut_norm'])

    # 3. LÓGICA DE NEGOCIO HISTÓRICA

    # Escenario 1: Match (Existen en A y B) -> Mantener datos con ESTADO REAL DE A
    df_match = df_b_mapped[df_b_mapped['rut_norm'].isin(ruts_a)].copy()
    #df_match['estado'] = df_match['rut_norm'].map(df_a_mapped.set_index('rut_norm')['estado'])
    mapa_estados = df_a_mapped.drop_duplicates(subset=['rut_norm'], keep='first').set_index('rut_norm')['estado']
    df_match['estado'] = df_match['rut_norm'].map(mapa_estados)
    df_match = df_match[cols_finales]

    # Escenario 2: Nuevos en B (No están en A) -> Alumnos antiguos que se deben Insertar como INACTIVOS
    df_historicos_nuevos = df_b_mapped[~df_b_mapped['rut_norm'].isin(ruts_a)].copy()
    df_historicos_nuevos['estado'] = 'Inactivo'
    df_historicos_nuevos = df_historicos_nuevos[cols_finales]

    # Escenario 3: Solo en A (No están en B) -> Ignorar para carga pero guardar en auditoría
    df_ignorados = df_a_mapped[~df_a_mapped['rut_norm'].isin(ruts_b)][cols_finales].copy()

    # 4. Generación de Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hojas = [
        ('Nuevos a Insertar (Inactivos)', df_historicos_nuevos, '842323'), # Prioritario para carga
        ('Coincidencias (Mantiene Estado A)', df_match, '1F4E78'),
        ('Solo en BD Actual (Ignorados)', df_ignorados, '595959')
    ]

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    for sheet_title, df_data, fill_color in hojas:
        ws = wb.create_sheet(title=sheet_title)
        ws.append(cols_finales)
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        
        for col_idx in range(1, len(cols_finales) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        fill_even = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for row_idx, row_values in enumerate(df_data.values, start=2):
            ws.append(list(row_values))
            row_fill = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx in range(1, len(cols_finales) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = row_fill
                c.border = thin_border
                c.font = Font(name='Calibri', size=10)
                col_name = cols_finales[col_idx - 1]
                if col_name in ['rut', 'genero', 'fecha_nacimiento', 'estado']:
                    c.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    c.alignment = Alignment(horizontal='left', vertical='center')
                    
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.views.sheetView[0].showGridLines = True

    try:
        wb.save(FILE_OUTPUT)
        print(f"\n✅ Archivo generado exitosamente: {FILE_OUTPUT}")
    except PermissionError:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        alt_output = FILE_OUTPUT.replace(".xlsx", f"_{timestamp}.xlsx")
        print(f"\n⚠️ Archivo en uso. Guardando en: {alt_output}")
        wb.save(alt_output)

    print(f"  - Nuevos a Insertar (Inactivos): {len(df_historicos_nuevos)} registros")
    print(f"  - Coincidencias (Mantiene Estado A): {len(df_match)} registros")
    print(f"  - Solo en BD Actual (Ignorados): {len(df_ignorados)} registros")

if __name__ == '__main__':
    procesar_comparacion_historico()